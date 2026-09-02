"""Excel (.xlsx) report export for ROVER extractions.

Two sheets, dynamic columns (schema-on-read — works for any document type):
  * Documents — one row per document (header fields + doc_id/job_id/document_name/status)
  * Products  — one row per line item (item fields + doc_id/document_name/job_id)

`workbook_bytes()` builds the whole store; `one_doc_bytes(doc)` builds a single
document (key/value header sheet + its products). Returns raw .xlsx bytes.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import numeric

from . import store

_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="CC785C")   # ROVER terracotta
_META = ["doc_id", "job_id", "document_name", "needs_review"]


def _num(v):
    # An amount printed with its currency used to export as a blank cell.
    return numeric.to_float(v)


def _style_header(ws):
    for c in ws[1]:
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_rows(ws, rows, lead=()):
    """Write dict rows with a stable column order: lead columns first, then the
    union of remaining keys (sorted). Numbers written as numbers."""
    if not rows:
        ws.append(["(no data)"])
        return
    keys = list(lead)
    for r in rows:
        for k in r.keys():
            if k not in keys and not k.startswith("_"):
                keys.append(k)
    ws.append(keys)
    _style_header(ws)
    for r in rows:
        out = []
        for k in keys:
            v = r.get(k)
            n = _num(v)
            out.append(n if (n is not None and not isinstance(v, bool)) else v)
        ws.append(out)
    for i, k in enumerate(keys, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(len(k) + 2, 12), 40)


def workbook_bytes() -> bytes:
    """Whole-store report: Documents sheet + Products sheet."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Documents"
    _write_rows(ws1, store.overall_rows(), lead=["document_name", "doc_id", "job_id", "needs_review"])
    ws2 = wb.create_sheet("Products")
    _write_rows(ws2, store.product_rows(), lead=["document_name", "doc_id", "no", "hs_code"])
    return _dump(wb)


def one_doc_bytes(doc: dict) -> bytes:
    """Single-document report: a key/value header sheet + its products."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Fields"
    ws1.append(["Field", "Value", "Read by", "Confidence", "Evidence"])
    _style_header(ws1)
    record = doc.get("record") or {}
    values = doc.get("values") or {}
    if record:
        for col, cell in record.items():
            if not isinstance(cell, dict):
                cell = {"value": cell}
            v = cell.get("value")
            if v in (None, ""):
                continue
            n = _num(v)
            ws1.append([col, n if n is not None else v, (cell.get("model") or "").split("/")[-1],
                        cell.get("confidence"), cell.get("source")])
    else:
        for col, v in values.items():
            if v in (None, ""):
                continue
            n = _num(v)
            ws1.append([col, n if n is not None else v, "", "", ""])
    for col, w in (("A", 26), ("B", 24), ("C", 14), ("D", 12), ("E", 46)):
        ws1.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Products")
    items = doc.get("items") or []
    _write_rows(ws2, [dict(it) for it in items if isinstance(it, dict)], lead=["no", "hs_code"])
    return _dump(wb)


def _dump(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
